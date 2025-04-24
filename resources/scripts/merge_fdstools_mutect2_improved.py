import argparse
import pandas as pd
import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def merge_variant_callers(file_fdstools: str, file_mutect2: str) -> pd.DataFrame:
    try:
        df1 = pd.read_csv(file_fdstools, sep="\t")
        df2 = pd.read_csv(file_mutect2, sep="\t")
    except Exception as e:
        print(f"Error reading input files: {e}", file=sys.stderr)
        raise

    try:
        df1["variant"] = df1["sequence"]
        df2["variant"] = df2["EMPOP_Variant"]

        merged = pd.merge(df1, df2, on="variant", how="outer", suffixes=("_FDSTOOLS", "_MUTECT2"))
        merged["called_by_FDSTOOLS"] = ~merged["sequence"].isna()
        merged["called_by_MUTECT2"] = ~merged["EMPOP_Variant"].isna()

        def extract_position(seq):
            match = re.search(r"(\d+\.?\d*)", str(seq))
            return float(match.group(1)) if match else float('inf')

        merged["variant_position"] = merged["variant"].apply(extract_position)
        merged = merged.sort_values(by="variant_position").drop(columns=["variant_position"])

        front = ["variant", "called_by_FDSTOOLS", "called_by_MUTECT2"]
        other = [col for col in merged.columns if col not in front]
        return merged[front + other]

    except Exception as e:
        print(f"Error processing and merging data: {e}", file=sys.stderr)
        raise

def apply_excel_styles(excel_path: str):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        # Define fills
        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # D to M
        fill_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")   # N to X
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # False flags

        header = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for idx, cell in enumerate(row):
                col_name = header[idx]

                # Fill column D to M (index 3 to 12)
                if 3 <= idx <= 12:
                    cell.fill = fill_green

                # Fill column N to X (index 13 to 23)
                elif 13 <= idx <= 23:
                    cell.fill = fill_blue

                # Red fill for False flags
                if col_name == "called_by_FDSTOOLS" and cell.value is False:
                    cell.fill = fill_green
                elif col_name == "called_by_MUTECT2" and cell.value is False:
                    cell.fill = fill_red

                if col_name.endswith("_MUTECT2") or col_name in ["EMPOP_Variant", "VariantLevel"]:
                    cell.fill = fill_blue
                elif col_name.endswith("_FDSTOOLS") or col_name in ["sequence", "total"]:
                    cell.fill = fill_green

                if col_name == "called_by_FDSTOOLS" and cell.value is False:
                    cell.fill = fill_red
                elif col_name == "called_by_MUTECT2" and cell.value is False:
                    cell.fill = fill_red
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get column name (e.g., 'A', 'B', etc.)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 1)
            ws.column_dimensions[column].width = adjusted_width


        wb.save(excel_path)
    except Exception as e:
        print(f"Error styling Excel file: {e}", file=sys.stderr)
        raise

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Merge mitochondrial variant calls from FDSTOOLS and MUTECT2.")
    parser.add_argument("caller1", help="Path to the FDSTOOLS file (TSV format, with 'sequence' column).")
    parser.add_argument("caller2", help="Path to the MUTECT2 file (TSV format, with 'EMPOP_Variant' column).")
    parser.add_argument("output_file", help="Path to save the merged output (XLSX format).")

    args = parser.parse_args()

    try:
        df_merged = merge_variant_callers(args.caller1, args.caller2)
        df_merged.drop(columns=["ID","is_noise_or_low_frq"], errors="ignore", inplace=True)
        df_merged.rename(columns={
            "interpolated_total_coverage": "interp_total",
            "total_wo_noise_or_low_frq": "total_clean",
            "variant_frequency_wo_noise_or_low_frq": "variant_frequency_clean"
        }, inplace=True)
        if "VariantLevel" in df_merged.columns:
            df_merged["VariantLevel"] = (df_merged["VariantLevel"] * 100).round(2)
        df_merged.to_excel(args.output_file, index=False, engine="openpyxl")
        print(f"Merged Excel file saved to: {args.output_file}")
    except Exception:
        print("Merging failed. Please check your input files and formats.", file=sys.stderr)
        sys.exit(1)

    try:
        apply_excel_styles(args.output_file)
        print("Excel styling completed successfully.")
    except Exception:
        print("Styling failed. The Excel file was created, but formatting could not be applied.", file=sys.stderr)
        sys.exit(1)
